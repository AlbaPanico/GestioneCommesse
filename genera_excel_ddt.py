import sys
import os
import json
import datetime
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re

LOG_PATH = os.path.join(os.path.dirname(__file__), "log_ddt.txt")

def log(msg):
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{now}] {msg}\n"
    try:
        with open(LOG_PATH, "a", encoding="utf-8") as f:
            f.write(line)
    except Exception:
        pass
    print(line.strip())

def log_non_eseguito(msg):
    log(f"NON ESEGUITO: {msg}")

def format_date_short(date_str):
    if not date_str:
        return ""
    date_str = str(date_str).replace("-", "/")
    try:
        g, m, a = date_str.split("/")
        a = a[-2:]
        return f"{g}/{m}/{a}"
    except Exception:
        return str(date_str)[-8:]

def estrai_nome_commessa_da_underscores(raw):
    """
    Prende la parte tra il 1° e il 2° underscore.
    Se non ci sono almeno 2 underscore, restituisce la stringa ripulita.
    """
    if not raw:
        return ""
    parts = str(raw).split("_")
    return parts[1].strip() if len(parts) >= 2 else str(raw).strip()

def _parse_float_safe(val, default=""):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        try:
            return float(val)
        except Exception:
            return default
    try:
        s = str(val).strip().replace(",", ".")
        if s == "":
            return default
        return float(s)
    except Exception:
        return default

# ───────────────────────────────────────────────────────────────────────────────
# Normalizza il codice commessa:
# - "C8888-11" resta così
# - "BBB_xxx_P_C8888-11" -> "C8888-11"
# ───────────────────────────────────────────────────────────────────────────────
def normalizza_codice_commessa(raw):
    if not raw:
        return ""
    s = str(raw).strip()
    if "_" in s:
        s = s.split("_")[-1].strip()
    return s

# Accetta "dd/mm/yyyy", "dd-mm-yyyy", "yyyy-mm-dd"
def parse_date_components_any(s):
    if not s:
        return None
    txt = str(s).strip()
    m = re.search(r'^\s*(\d{2})[\/\-](\d{2})[\/\-](\d{4})\s*$', txt)
    if m:
        dd, mm, yyyy = m.group(1), m.group(2), m.group(3)
        return dd, mm, yyyy
    m = re.search(r'^\s*(\d{4})[\/\-](\d{2})[\/\-](\d{2})\s*$', txt)
    if m:
        yyyy, mm, dd = m.group(1), m.group(2), m.group(3)
        return dd, mm, yyyy
    m = re.search(r'(\d{2}).*(\d{2}).*(\d{4})', txt)
    if m:
        return m.group(1), m.group(2), m.group(3)
    return None

def build_expected_filenames(numero_ddt, codice_commessa_norm, codice_commessa_raw, dd_mm_yyyy):
    """
    Genera possibili nomi attesi usando sia il codice normalizzato sia il raw.
    Varianti '-' o '_' nella data, estensione .pdf/.PDF.
    """
    out = []
    if not (numero_ddt and dd_mm_yyyy):
        return out
    dd, mm, yyyy = dd_mm_yyyy
    # num a 4 cifre (se numero_ddt contiene già stringhe, left pad)
    num4 = f"{int(str(numero_ddt)) :04d}" if str(numero_ddt).isdigit() else f"{str(numero_ddt):0>4}"

    codes = []
    if codice_commessa_raw:
        codes.append(codice_commessa_raw)
    if codice_commessa_norm and codice_commessa_norm not in codes:
        codes.append(codice_commessa_norm)

    for code in codes:
        bases = [
            f"DDT_{num4}W_{code}_{dd}-{mm}-{yyyy}",
            f"DDT_{num4}W_{code}_{dd}_{mm}_{yyyy}",
        ]
        for b in bases:
            out.append(b + ".pdf")
            out.append(b + ".PDF")
    return out

def match_ddt_files(folder, codice_norm, codice_raw):
    """
    Scansione robusta:
    - pattern 1: RAW esatto (W_<raw>_dd-mm-yyyy)
    - pattern 2: qualsiasi prefisso prima del codice NORMALIZZATO (W_.*<norm>_dd-mm-yyyy)
    Estensione .pdf/.PDF, case-insensitive, '-' o '_' nella data, permetti suffissi dopo la data.
    Ritorna: liste [(num, fname, dd/mm/yyyy)] per ENTRATA (W) e USCITA (T).
    """
    pats_entrata = []
    pats_uscita = []
    flags = re.IGNORECASE

    if codice_raw:
        pats_entrata.append(re.compile(
            rf'^DDT_(\d{{4}})W_{re.escape(codice_raw)}_(\d{{2}})[\-_](\d{{2}})[\-_](\d{{4}})(?:\D.*)?\.pdf$',
            flags
        ))
        pats_uscita.append(re.compile(
            rf'^DDT_(\d{{4}})T_{re.escape(codice_raw)}_(\d{{2}})[\-_](\d{{2}})[\-_](\d{{4}})(?:\D.*)?\.pdf$',
            flags
        ))

    if codice_norm:
        pats_entrata.append(re.compile(
            rf'^DDT_(\d{{4}})W_.*{re.escape(codice_norm)}_(\d{{2}})[\-_](\d{{2}})[\-_](\d{{4}})(?:\D.*)?\.pdf$',
            flags
        ))
        pats_uscita.append(re.compile(
            rf'^DDT_(\d{{4}})T_.*{re.escape(codice_norm)}_(\d{{2}})[\-_](\d{{2}})[\-_](\d{{4}})(?:\D.*)?\.pdf$',
            flags
        ))

    entrata, uscita = [], []
    try:
        files = os.listdir(folder)
    except Exception as e:
        log(f"ERRORE os.listdir: {e}")
        return entrata, uscita

    preview = ", ".join(files[:20])
    log(f"DEBUG MATERIALI contiene ({len(files)} file): {preview}{'...' if len(files)>20 else ''}")

    for fname in files:
        for pat in pats_entrata:
            m = pat.match(fname)
            if m:
                num = int(m.group(1))
                g, mth, a = m.group(2), m.group(3), m.group(4)
                entrata.append((num, fname, f"{g}/{mth}/{a}"))
                break  # evita doppioni
        for pat in pats_uscita:
            m = pat.match(fname)
            if m:
                num = int(m.group(1))
                g, mth, a = m.group(2), m.group(3), m.group(4)
                uscita.append((num, fname, f"{g}/{mth}/{a}"))
                break
    return entrata, uscita

def main():
    if len(sys.argv) != 3:
        print("Usage: genera_excel_ddt.py dati.json basePath", file=sys.stderr)
        sys.exit(1)

    dati_path = sys.argv[1]
    base_path = sys.argv[2]
    template_path = os.path.join(base_path, "Template", "DDT_Work.xlsx")

    # --- Leggi dati JSON ---
    with open(dati_path, "r", encoding="utf-8") as f:
        dati = json.load(f)

    # Log ingresso
    log(f"DEBUG DATI keys={list(dati.keys())}")
    log(f"DEBUG nomeCommessa='{dati.get('nomeCommessa','')}' descrizione='{dati.get('descrizione','')}'")

    # Nome commessa (visivo colonna D)
    sorgente_descr = dati.get("descrizione") or dati.get("nomeCommessa") or ""
    nome_commessa = estrai_nome_commessa_da_underscores(sorgente_descr)
    log(f"DEBUG estrazione: sorgente='{sorgente_descr}' -> estratto='{nome_commessa}'")

    # Cartella MATERIALI
    folder_materiali = ""
    if dati.get("percorsoPdf", ""):
        folder_materiali = os.path.dirname(dati["percorsoPdf"])
    elif dati.get("folderPath", ""):
        folder_materiali = os.path.join(dati["folderPath"], "MATERIALI")
    folder_materiali = os.path.normpath(folder_materiali)
    log(f"DEBUG folder_materiali='{folder_materiali}' esiste={os.path.isdir(folder_materiali)}")

    # Codici commessa
    codice_commessa_raw = dati.get("codiceCommessa", "")
    codice_commessa_norm = normalizza_codice_commessa(codice_commessa_raw)
    log(f"DEBUG codice_commessa_norm='{codice_commessa_norm}' (raw='{codice_commessa_raw}')")

    # Dati DDT da JSON
    numero_ddt_json = (str(dati.get("numeroDdt", "")).strip() or str(dati.get("numeroDdtEntrata","")).strip())
    data_ddt_json   = (str(dati.get("dataDdt", "")).strip() or str(dati.get("dataDdtEntrata","")).strip())
    num_ddt_uscita_json = (str(dati.get("nsDdt", "")).strip() or str(dati.get("numeroDdtUscita","")).strip())
    data_ddt_uscita_json = (str(dati.get("del", "")).strip() or str(dati.get("dataDdtUscita","")).strip())

    # ------------ TROVA PDF ENTRATA (OBBLIGATORIO) ------------
    link_pdf_entrata = ""
    numero_ddt_entrata = ""
    data_ddt_entrata = ""

    if os.path.isdir(folder_materiali):
        # 1) Nomi attesi da numero + data, sia con raw che con norm
        date_cmp = parse_date_components_any(data_ddt_json)
        if date_cmp:
            for cand in build_expected_filenames(numero_ddt_json, codice_commessa_norm, codice_commessa_raw, date_cmp):
                cand_path = os.path.join(folder_materiali, cand)
                if os.path.exists(cand_path):
                    link_pdf_entrata = cand_path
                    numero_ddt_entrata = re.search(r'DDT_(\d{4})W_', cand, re.IGNORECASE).group(1) + "W"
                    dd, mm, yyyy = date_cmp
                    data_ddt_entrata = f"{dd}/{mm}/{yyyy}"
                    log(f"DEBUG match atteso ENTRATA: {cand}")
                    break

        # 2) Scansione regex permissiva (raw esatto oppure qualunque prefisso + code norm)
        if not link_pdf_entrata:
            pdf_entrata_candidates, pdf_uscita_candidates = match_ddt_files(
                folder_materiali, codice_commessa_norm, codice_commessa_raw
            )
            if pdf_entrata_candidates:
                pdf_entrata_candidates.sort(reverse=True)  # prendo il numero più alto
                num_w, fname_w, data_w = pdf_entrata_candidates[0]
                link_pdf_entrata = os.path.join(folder_materiali, fname_w)
                numero_ddt_entrata = f"{str(num_w).zfill(4)}W"
                data_ddt_entrata = data_w
                log(f"DEBUG match regex ENTRATA: {fname_w}")

    # Obbligatorio: se non trovo ENTRATA, esco
    if not link_pdf_entrata or not os.path.exists(link_pdf_entrata):
        log_non_eseguito(f"Excel DDT NON aggiornato: DDT Entrata NON trovato (folder='{folder_materiali}')")
        return

    # ------------ TROVA PDF USCITA (OPZIONALE) ------------
    link_pdf_uscita = ""
    numero_ddt_uscita = ""
    data_ddt_uscita = ""

    if os.path.isdir(folder_materiali):
        # tentativo da JSON
        date_cmp_u = parse_date_components_any(data_ddt_uscita_json)
        if date_cmp_u and num_ddt_uscita_json:
            dd, mm, yyyy = date_cmp_u
            num4 = f"{int(str(num_ddt_uscita_json)) :04d}" if str(num_ddt_uscita_json).isdigit() else f"{str(num_ddt_uscita_json):0>4}"
            candidates = [
                f"DDT_{num4}T_{codice_commessa_raw}_{dd}-{mm}-{yyyy}.pdf",
                f"DDT_{num4}T_{codice_commessa_raw}_{dd}_{mm}_{yyyy}.pdf",
                f"DDT_{num4}T_{codice_commessa_norm}_{dd}-{mm}-{yyyy}.pdf",
                f"DDT_{num4}T_{codice_commessa_norm}_{dd}_{mm}_{yyyy}.pdf",
                f"DDT_{num4}T_{codice_commessa_raw}_{dd}-{mm}-{yyyy}.PDF",
                f"DDT_{num4}T_{codice_commessa_raw}_{dd}_{mm}_{yyyy}.PDF",
                f"DDT_{num4}T_{codice_commessa_norm}_{dd}-{mm}-{yyyy}.PDF",
                f"DDT_{num4}T_{codice_commessa_norm}_{dd}_{mm}_{yyyy}.PDF",
            ]
            for cand in candidates:
                p = os.path.join(folder_materiali, cand)
                if os.path.exists(p):
                    link_pdf_uscita = p
                    numero_ddt_uscita = f"{num4}T"
                    data_ddt_uscita = f"{dd}/{mm}/{yyyy}"
                    log(f"DEBUG match atteso USCITA: {cand}")
                    break

        # regex permissiva
        if not numero_ddt_uscita:
            ent, usc = match_ddt_files(folder_materiali, codice_commessa_norm, codice_commessa_raw)
            if usc:
                usc.sort(reverse=True)
                num_t, fname_t, data_t = usc[0]
                link_pdf_uscita = os.path.join(folder_materiali, fname_t)
                numero_ddt_uscita = f"{str(num_t).zfill(4)}T"
                data_ddt_uscita = data_t
                log(f"DEBUG match regex USCITA: {fname_t}")

    # ------------ PREPARA EXCEL ------------
    data_ddt_entrata_short = format_date_short(data_ddt_entrata)
    data_ddt_uscita_short = format_date_short(data_ddt_uscita)

    # Mese/anno per file e intestazione
    try:
        _, mese, anno = data_ddt_entrata.split("/")
    except Exception:
        now = datetime.datetime.now()
        mese = f"{now.month:02d}"
        anno = str(now.year)

    file_name = f"DDT_Work_{mese}_{anno}.xlsx"
    file_path = os.path.join(base_path, file_name)

    if not os.path.exists(file_path):
        if not os.path.exists(template_path):
            print("Template non trovato:", template_path, file=sys.stderr)
            sys.exit(2)
        shutil.copy(template_path, file_path)

    wb = load_workbook(file_path)
    ws = wb.active

    mesi_it = ["", "GENNAIO", "FEBBRAIO", "MARZO", "APRILE", "MAGGIO", "GIUGNO",
               "LUGLIO", "AGOSTO", "SETTEMBRE", "OTTOBRE", "NOVEMBRE", "DICEMBRE"]
    ws["A1"] = f"DDT Work Report {mesi_it[int(mese)]} {anno}"

       # Dedup: se esiste già una riga con lo stesso Numero DDT ENTRATA (col. B)
    # oppure lo stesso hyperlink del PDF (col. I), sovrascrivo quella riga
    existing_row = None
    try:
        maxr = ws.max_row or 5
        for r in range(5, maxr + 1):
            # match su Numero DDT (es. "0078W")
            valB = ws[f"B{r}"].value
            if valB and str(valB).strip() == str(numero_ddt_entrata).strip():
                existing_row = r
                break
            # match su hyperlink (target) in colonna I
            try:
                hl = ws[f"I{r}"].hyperlink
                target = getattr(hl, "target", None) if hl else None
                if target and os.path.normcase(target) == os.path.normcase(link_pdf_entrata):
                    existing_row = r
                    break
            except Exception:
                pass
    except Exception:
        pass

    if existing_row:
        row = existing_row
        log(f"DEBUG dedupe: sovrascrivo riga esistente {row} per DDT {numero_ddt_entrata}")
    else:
        # Prima riga libera
        row = 5
        while ws[f"A{row}"].value:
            row += 1
        log(f"DEBUG dedupe: nessuna riga esistente, scrivo in nuova riga {row}")


    # Nome visivo
    valore_nome = nome_commessa if nome_commessa else "(NO NOME)"

    # Prezzo vendita unitario
    prezzo_vendita = dati.get("prezzoVendita", 0)
    prezzo_vendita_num = _parse_float_safe(prezzo_vendita, default=0.0)
    log(f"DEBUG prezzoVendita ricevuto='{prezzo_vendita}' -> usato={prezzo_vendita_num}")

    # Scrivi riga
    ws[f"A{row}"] = data_ddt_entrata_short
    ws[f"B{row}"] = numero_ddt_entrata
    ws[f"C{row}"] = normalizza_codice_commessa(dati.get("codiceCommessa", ""))  # C8888-11 in chiaro
    ws[f"D{row}"] = valore_nome
    ws[f"E{row}"] = _parse_float_safe(dati.get("quantita"), default="")
    ws[f"F{row}"] = dati.get("colli", "")
    ws[f"G{row}"] = numero_ddt_uscita
    ws[f"H{row}"] = data_ddt_uscita_short

    cell = ws[f"I{row}"]; cell.value = "Apri"; cell.hyperlink = link_pdf_entrata; cell.style = "Hyperlink"

    ore = dati.get("oreLavorazione")
    ws[f"J{row}"] = _parse_float_safe(ore, default="")

    ws[f"K{row}"] = f'=IF(E{row}<>0,L{row}/E{row},"")'
    ws[f"L{row}"] = f'=M{row}*J{row}'
    ws[f"M{row}"] = ""
    ws[f"N{row}"] = prezzo_vendita_num

    # Allinea
    for col in "ABCDEFGHIJKLMN":
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")

    wb.save(file_path)
    log(f"OK: scritto su {file_path} (riga {row}) | NomeCommessa='{valore_nome}' | Codice='{normalizza_codice_commessa(dati.get('codiceCommessa',''))}' | PrezzoVendita={prezzo_vendita_num} | PDF_IN='{os.path.basename(link_pdf_entrata)}'")

if __name__ == "__main__":
    main()
